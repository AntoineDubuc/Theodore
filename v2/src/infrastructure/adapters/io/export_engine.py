#!/usr/bin/env python3
"""
Theodore v2 Export Engine Implementation

Comprehensive export engine supporting multiple formats with streaming capabilities.
"""

import asyncio
import os
import json
import csv
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any, Union
from pathlib import Path
import time

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import PieChart, BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from ....core.ports.io.export_engine_port import ExportEnginePort
from ....core.domain.models.company import CompanyData
from ....core.domain.models.export import (
    ExportResult, OutputConfig, Visualization, ExportFormat
)
from ...observability.logging import get_logger
from ...observability.metrics import get_metrics_collector

logger = get_logger(__name__)
metrics = get_metrics_collector()


class ExportError(Exception):
    """Export operation error"""
    pass


class UnsupportedFormatError(ExportError):
    """Unsupported export format error"""
    pass


class ExportEngine(ExportEnginePort):
    """
    Comprehensive export engine supporting multiple formats
    
    Supports:
    - CSV, JSON, Excel, PDF exports
    - Streaming for large datasets
    - Visualization integration
    - Format-specific optimizations
    """
    
    def __init__(self):
        self.formatters = {
            ExportFormat.CSV: CSVExporter(),
            ExportFormat.JSON: JSONExporter(),
            ExportFormat.EXCEL: ExcelExporter(),
            ExportFormat.PDF: PDFExporter(),
            ExportFormat.PARQUET: ParquetExporter(),
            ExportFormat.POWERBI: PowerBIExporter(),
            ExportFormat.TABLEAU: TableauExporter()
        }
        
    async def export(
        self,
        data: List[CompanyData],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export data using specified configuration"""
        
        start_time = time.time()
        
        try:
            # Get appropriate formatter
            formatter = self.formatters.get(config.format)
            if not formatter:
                available_formats = list(self.formatters.keys())
                raise UnsupportedFormatError(
                    f"Format '{config.format}' not supported. "
                    f"Available: {[f.value for f in available_formats]}"
                )
            
            # Check if formatter is available (dependencies installed)
            if not formatter.is_available():
                missing_deps = formatter.get_missing_dependencies()
                raise ExportError(
                    f"Cannot export to {config.format.value} format. "
                    f"Missing dependencies: {', '.join(missing_deps)}"
                )
            
            # Apply data transformations
            transformed_data = await self._transform_data(data, config)
            
            # Determine if streaming is needed
            should_stream = (
                config.stream_large_datasets and 
                len(transformed_data) > config.streaming_threshold
            )
            
            if should_stream:
                logger.info(
                    f"Using streaming export for {len(transformed_data)} records "
                    f"(threshold: {config.streaming_threshold})"
                )
                result = await self._stream_export(formatter, transformed_data, config)
            else:
                result = await self._standard_export(
                    formatter, transformed_data, config, visualizations
                )
            
            # Calculate processing time
            processing_time = time.time() - start_time
            result.processing_time_seconds = processing_time
            
            # Record metrics
            metrics.record_histogram(
                "export_processing_time_seconds",
                processing_time,
                tags={"format": config.format.value}
            )
            
            metrics.record_histogram(
                "export_file_size_bytes",
                result.file_size_bytes,
                tags={"format": config.format.value}
            )
            
            logger.info(
                f"Export completed successfully",
                extra={
                    "format": config.format.value,
                    "record_count": result.record_count,
                    "file_size_mb": result.file_size_mb,
                    "processing_time": processing_time,
                    "streaming_used": result.streaming_used
                }
            )
            
            return result
            
        except Exception as e:
            logger.error(f"Export failed: {e}", exc_info=True)
            raise ExportError(f"Export operation failed: {e}") from e
    
    async def validate_export_config(self, config: OutputConfig) -> bool:
        """Validate export configuration"""
        try:
            # Check if format is supported
            if config.format not in self.formatters:
                return False
            
            # Check if formatter is available
            formatter = self.formatters[config.format]
            if not formatter.is_available():
                return False
            
            # Validate output path
            output_path = config.file_path
            if not output_path.parent.exists():
                output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Validate streaming configuration
            if config.streaming_threshold <= 0:
                return False
            
            if config.chunk_size <= 0:
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Config validation failed: {e}")
            return False
    
    async def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats"""
        supported = []
        for format_enum, formatter in self.formatters.items():
            if formatter.is_available():
                supported.append(format_enum.value)
        return supported
    
    async def estimate_export_size(
        self,
        data: List[CompanyData],
        config: OutputConfig
    ) -> int:
        """Estimate export file size in bytes"""
        if not data:
            return 0
        
        # Get sample record size
        sample_record = data[0]
        sample_size = len(json.dumps(sample_record.dict()))
        
        # Estimate based on format
        format_multipliers = {
            ExportFormat.JSON: 1.0,
            ExportFormat.CSV: 0.7,
            ExportFormat.EXCEL: 1.2,
            ExportFormat.PDF: 2.0,
            ExportFormat.PARQUET: 0.3,
            ExportFormat.POWERBI: 1.1,
            ExportFormat.TABLEAU: 0.8
        }
        
        multiplier = format_multipliers.get(config.format, 1.0)
        estimated_size = int(sample_size * len(data) * multiplier)
        
        # Account for compression
        if config.compress:
            estimated_size = int(estimated_size * 0.6)  # Assume 40% compression
        
        return estimated_size
    
    async def cleanup_temp_files(self, export_result: ExportResult) -> None:
        """Clean up temporary files created during export"""
        # For now, we don't create temp files
        # This could be implemented for complex multi-file exports
        pass
    
    async def _transform_data(
        self, 
        data: List[CompanyData], 
        config: OutputConfig
    ) -> List[Dict[str, Any]]:
        """Transform company data for export"""
        transformed = []
        
        for company in data:
            # Convert to dictionary
            company_dict = company.dict()
            
            # Apply field selection if specified
            include_fields = getattr(config, 'include_fields', [])
            exclude_fields = getattr(config, 'exclude_fields', [])
            
            if include_fields:
                company_dict = {
                    k: v for k, v in company_dict.items() 
                    if k in include_fields
                }
            
            if exclude_fields:
                company_dict = {
                    k: v for k, v in company_dict.items() 
                    if k not in exclude_fields
                }
            
            # Apply custom column mappings
            custom_columns = getattr(config, 'custom_columns', {})
            for new_name, field_path in custom_columns.items():
                # Simple field mapping (could be enhanced for nested fields)
                if field_path in company_dict:
                    company_dict[new_name] = company_dict[field_path]
            
            transformed.append(company_dict)
        
        return transformed
    
    async def _standard_export(
        self,
        formatter: 'ExportFormatter',
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]]
    ) -> ExportResult:
        """Handle standard (non-streaming) export"""
        return await formatter.export(data, config, visualizations)
    
    async def _stream_export(
        self,
        formatter: 'ExportFormatter',
        data: List[Dict[str, Any]],
        config: OutputConfig
    ) -> ExportResult:
        """Handle streaming export for large datasets"""
        chunk_size = config.chunk_size
        total_chunks = (len(data) + chunk_size - 1) // chunk_size
        
        # Initialize streaming export
        export_context = await formatter.initialize_streaming_export(config)
        
        try:
            for i in range(0, len(data), chunk_size):
                chunk = data[i:i + chunk_size]
                chunk_number = (i // chunk_size) + 1
                
                # Export chunk
                await formatter.export_chunk(chunk, export_context)
                
                # Report progress
                progress = (chunk_number / total_chunks) * 100
                logger.debug(
                    f"Streaming export progress: {progress:.1f}% "
                    f"({chunk_number}/{total_chunks} chunks)"
                )
                
                # Allow other tasks to run
                await asyncio.sleep(0)
            
            # Finalize streaming export
            result = await formatter.finalize_streaming_export(export_context)
            result.streaming_used = True
            result.chunks_processed = total_chunks
            
            return result
            
        except Exception as e:
            # Cleanup on failure
            await formatter.cleanup_failed_export(export_context)
            raise ExportError(f"Streaming export failed: {e}") from e


class ExportFormatter:
    """Base class for format-specific exporters"""
    
    def is_available(self) -> bool:
        """Check if formatter dependencies are available"""
        return True
    
    def get_missing_dependencies(self) -> List[str]:
        """Get list of missing dependencies"""
        return []
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export data in specific format"""
        raise NotImplementedError
    
    async def initialize_streaming_export(self, config: OutputConfig) -> Dict[str, Any]:
        """Initialize streaming export context"""
        return {}
    
    async def export_chunk(
        self, 
        chunk: List[Dict[str, Any]], 
        context: Dict[str, Any]
    ) -> None:
        """Export a data chunk in streaming mode"""
        raise NotImplementedError
    
    async def finalize_streaming_export(self, context: Dict[str, Any]) -> ExportResult:
        """Finalize streaming export and return result"""
        raise NotImplementedError
    
    async def cleanup_failed_export(self, context: Dict[str, Any]) -> None:
        """Clean up after failed export"""
        pass


class CSVExporter(ExportFormatter):
    """CSV format exporter"""
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export to CSV format"""
        
        if not data:
            # Create empty CSV file
            with open(config.file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([])  # Empty header
            
            return ExportResult(
                file_path=config.file_path,
                format=ExportFormat.CSV,
                record_count=0,
                file_size_bytes=os.path.getsize(config.file_path),
                columns=[]
            )
        
        # Get column names from first record
        columns = list(data[0].keys())
        
        # Write CSV file
        with open(config.file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)
        
        return ExportResult(
            file_path=config.file_path,
            format=ExportFormat.CSV,
            record_count=len(data),
            file_size_bytes=os.path.getsize(config.file_path),
            columns=columns
        )
    
    async def initialize_streaming_export(self, config: OutputConfig) -> Dict[str, Any]:
        """Initialize CSV streaming export"""
        file_handle = open(config.file_path, 'w', newline='', encoding='utf-8')
        return {
            'file_handle': file_handle,
            'writer': None,
            'columns_written': False,
            'record_count': 0
        }
    
    async def export_chunk(
        self, 
        chunk: List[Dict[str, Any]], 
        context: Dict[str, Any]
    ) -> None:
        """Export CSV chunk"""
        if not chunk:
            return
        
        file_handle = context['file_handle']
        
        # Initialize writer on first chunk
        if context['writer'] is None:
            columns = list(chunk[0].keys())
            writer = csv.DictWriter(file_handle, fieldnames=columns)
            context['writer'] = writer
            context['columns'] = columns
            
            # Write header
            writer.writeheader()
            context['columns_written'] = True
        
        # Write chunk data
        writer = context['writer']
        writer.writerows(chunk)
        context['record_count'] += len(chunk)
    
    async def finalize_streaming_export(self, context: Dict[str, Any]) -> ExportResult:
        """Finalize CSV streaming export"""
        file_handle = context['file_handle']
        file_handle.close()
        
        file_path = Path(file_handle.name)
        return ExportResult(
            file_path=file_path,
            format=ExportFormat.CSV,
            record_count=context['record_count'],
            file_size_bytes=os.path.getsize(file_path),
            columns=context.get('columns', [])
        )
    
    async def cleanup_failed_export(self, context: Dict[str, Any]) -> None:
        """Cleanup failed CSV export"""
        if 'file_handle' in context:
            try:
                context['file_handle'].close()
            except:
                pass


class JSONExporter(ExportFormatter):
    """JSON format exporter"""
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export to JSON format"""
        
        export_data = {
            'companies': data,
            'metadata': {
                'export_timestamp': datetime.now(timezone.utc).isoformat(),
                'record_count': len(data),
                'format': 'json'
            }
        }
        
        # Include visualizations if provided
        if visualizations:
            export_data['visualizations'] = [v.dict() for v in visualizations]
        
        # Write JSON file
        with open(config.file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
        
        columns = list(data[0].keys()) if data else []
        
        return ExportResult(
            file_path=config.file_path,
            format=ExportFormat.JSON,
            record_count=len(data),
            file_size_bytes=os.path.getsize(config.file_path),
            columns=columns
        )


class ExcelExporter(ExportFormatter):
    """Excel format exporter with advanced features"""
    
    def is_available(self) -> bool:
        return HAS_OPENPYXL
    
    def get_missing_dependencies(self) -> List[str]:
        deps = []
        if not HAS_OPENPYXL:
            deps.append("openpyxl")
        return deps
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export to Excel with multiple sheets and formatting"""
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        workbook = Workbook()
        
        # Main data sheet
        companies_sheet = workbook.active
        companies_sheet.title = "Companies"
        
        if data:
            columns = list(data[0].keys())
            
            # Write headers
            for col_num, column in enumerate(columns, 1):
                cell = companies_sheet.cell(row=1, column=col_num, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_num, record in enumerate(data, 2):
                for col_num, column in enumerate(columns, 1):
                    value = record.get(column, '')
                    companies_sheet.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in companies_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                companies_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        summary_data = [
            ["Total Companies", len(data)],
            ["Export Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")],
            ["Export Format", "Excel (.xlsx)"]
        ]
        
        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_sheet.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            summary_sheet.cell(row=row_num, column=2, value=value)
        
        # Add visualizations sheet if provided
        if visualizations:
            viz_sheet = workbook.create_sheet("Charts")
            viz_sheet.cell(row=1, column=1, value="Visualizations").font = Font(bold=True, size=14)
            
            for i, viz in enumerate(visualizations):
                row = i + 3
                viz_sheet.cell(row=row, column=1, value=f"{viz.title}")
                viz_sheet.cell(row=row, column=2, value=f"Type: {viz.type.value}")
        
        # Save workbook
        workbook.save(config.file_path)
        
        sheets = [sheet.title for sheet in workbook.worksheets]
        columns = list(data[0].keys()) if data else []
        
        return ExportResult(
            file_path=config.file_path,
            format=ExportFormat.EXCEL,
            record_count=len(data),
            file_size_bytes=os.path.getsize(config.file_path),
            columns=columns,
            sheets=sheets
        )


class PDFExporter(ExportFormatter):
    """PDF format exporter"""
    
    def is_available(self) -> bool:
        return HAS_REPORTLAB
    
    def get_missing_dependencies(self) -> List[str]:
        deps = []
        if not HAS_REPORTLAB:
            deps.append("reportlab")
        return deps
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export to PDF format"""
        
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        doc = SimpleDocTemplate(str(config.file_path), pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph("Company Data Export Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Summary
        summary_text = f"""
        <para>
        <b>Export Summary:</b><br/>
        Total Companies: {len(data)}<br/>
        Export Date: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")}<br/>
        Format: PDF Report
        </para>
        """
        summary = Paragraph(summary_text, styles['Normal'])
        story.append(summary)
        story.append(Spacer(1, 12))
        
        # Data table
        if data:
            # Prepare table data
            columns = list(data[0].keys())
            # Limit columns for PDF readability
            display_columns = columns[:6] if len(columns) > 6 else columns
            
            table_data = [display_columns]  # Header row
            
            for record in data[:50]:  # Limit rows for PDF
                row = [str(record.get(col, '')) for col in display_columns]
                # Truncate long values
                row = [val[:30] + '...' if len(val) > 30 else val for val in row]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
            if len(data) > 50:
                note = Paragraph(
                    f"<i>Note: Showing first 50 of {len(data)} companies. "
                    "Full data available in other export formats.</i>",
                    styles['Italic']
                )
                story.append(Spacer(1, 12))
                story.append(note)
        
        # Build PDF
        doc.build(story)
        
        columns = list(data[0].keys()) if data else []
        
        return ExportResult(
            file_path=config.file_path,
            format=ExportFormat.PDF,
            record_count=len(data),
            file_size_bytes=os.path.getsize(config.file_path),
            columns=columns
        )


class ParquetExporter(ExportFormatter):
    """Parquet format exporter (requires pandas and pyarrow)"""
    
    def is_available(self) -> bool:
        try:
            import pandas as pd
            import pyarrow
            return True
        except ImportError:
            return False
    
    def get_missing_dependencies(self) -> List[str]:
        deps = []
        try:
            import pandas as pd
        except ImportError:
            deps.append("pandas")
        
        try:
            import pyarrow
        except ImportError:
            deps.append("pyarrow")
        
        return deps
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export to Parquet format"""
        
        import pandas as pd
        
        if not data:
            # Create empty parquet file
            df = pd.DataFrame()
            df.to_parquet(config.file_path)
            
            return ExportResult(
                file_path=config.file_path,
                format=ExportFormat.PARQUET,
                record_count=0,
                file_size_bytes=os.path.getsize(config.file_path),
                columns=[]
            )
        
        # Convert to DataFrame and save as Parquet
        df = pd.DataFrame(data)
        df.to_parquet(config.file_path, index=False)
        
        return ExportResult(
            file_path=config.file_path,
            format=ExportFormat.PARQUET,
            record_count=len(data),
            file_size_bytes=os.path.getsize(config.file_path),
            columns=list(df.columns)
        )


class PowerBIExporter(ExportFormatter):
    """PowerBI-optimized exporter"""
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export optimized for PowerBI import"""
        
        # For now, export as Excel which PowerBI can import
        # In future, could create .pbix files directly
        excel_exporter = ExcelExporter()
        if not excel_exporter.is_available():
            # Fallback to CSV
            csv_exporter = CSVExporter()
            return await csv_exporter.export(data, config, visualizations)
        
        result = await excel_exporter.export(data, config, visualizations)
        result.format = ExportFormat.POWERBI
        return result


class TableauExporter(ExportFormatter):
    """Tableau-optimized exporter"""
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        config: OutputConfig,
        visualizations: Optional[List[Visualization]] = None
    ) -> ExportResult:
        """Export optimized for Tableau import"""
        
        # For now, export as CSV which Tableau can import easily
        # In future, could create .hyper files directly
        csv_exporter = CSVExporter()
        result = await csv_exporter.export(data, config, visualizations)
        result.format = ExportFormat.TABLEAU
        return result